import openpyxl
import texttable
from datetime import date
from telebot.types import Message
from telebot.formatting import hbold, hcode

from bot import bot
from filters import CommandWeek, CommandToday, CommandTomorrow
from config import SHEDULE_TABLE


shedule_table = openpyxl.load_workbook(SHEDULE_TABLE, read_only=True)
shedule_week1 = shedule_table['неделя 1']
shedule_week2 = shedule_table['неделя 2']


# 0 - первая, 1 - вторая
def today_week():
    return 0 if date.today().isocalendar().week % 2 else 1


@bot.message_handler(**CommandWeek)
def week_handler(msg: Message):
    n = today_week() + 1
    bot.send_message(msg.chat.id, f"Сегодня неделя {hbold('№' + str(n))}")


@bot.message_handler(**CommandToday)
def today_handler(msg: Message):
    shedule = shedule_week2 if today_week() else shedule_week1
    column = 3 + date.today().weekday()

    ttable = texttable.Texttable()
    ttable.set_cols_align(["c", "c", "l"])

    for row in range(2, shedule.max_row + 1):
        start_t = shedule.cell(row, 1).value.strftime('%H:%M')
        end_t = shedule.cell(row, 2).value.strftime('%H:%M')
        subject = shedule.cell(row, column).value

        if not subject:
            subject = str()

        ttable.add_row([start_t, end_t, subject])

    bot.send_message(msg.chat.id, hcode(ttable.draw()))


@bot.message_handler(**CommandTomorrow)
def tomorrow_handler(msg: Message):
    pass


__all__ = [
    'week_handler',
    'today_handler',
    'tomorrow_handler',
]
